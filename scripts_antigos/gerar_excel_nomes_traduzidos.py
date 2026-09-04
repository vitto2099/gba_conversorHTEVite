from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TRANSLATIONS = {
    "nutzenbewertungsverfahren": "benefit assessment procedure",
    "zum": "for the",
    "wirkstoff": "active substance",
    "erneute": "renewed",
    "neubewertung": "reassessment",
    "chronische": "chronic",
    "chronischer": "chronic",
    "chronisch": "chronic",
    "diabetes": "diabetes",
    "pulmonal": "pulmonary",
    "schizophrenie": "schizophrenia",
    "multiresistente": "multidrug resistant",
    "multiresistenter": "multidrug resistant",
    "störung": "disorder",
    "stoerung": "disorder",
    "überschuss": "excess",
    "ueberschuss": "excess",
    "überaktive": "overactive",
    "ueberaktive": "overactive",
    "hiv": "HIV",
    "infektion": "infection",
    "infektionen": "infections",
    "prostatakarzinom": "prostate cancer",
    "schilddruesenkarzinom": "thyroid cancer",
    "schilddrüsenkarzinom": "thyroid cancer",
    "colitis": "colitis",
    "copd": "COPD",
    "neues": "new",
    "anwendungsgebiet": "therapeutic indication",
    "larven": "larvae",
    "lebende": "live",
    "von": "of",
    "aus": "from",
    "alfa": "alpha",
}


def load_translation_overrides(path: Path | None) -> dict[str, str]:
    """Load exact-name or token translation overrides from a CSV file."""
    if not path:
        return {}

    overrides: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"original", "english"}
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(
                f"CSV de traduções deve conter as colunas: {', '.join(sorted(required))}"
            )

        for row in reader:
            original = (row.get("original") or "").strip()
            english = (row.get("english") or "").strip()
            if original and english:
                overrides[normalize_key(original)] = english

    return overrides


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("_", " ").strip()).casefold()


def split_name(name: str) -> tuple[str, str]:
    path_name = Path(name)
    if path_name.suffix:
        return path_name.stem, path_name.suffix
    return name, ""


def translate_name(name: str, translations: dict[str, str]) -> str:
    stem, suffix = split_name(name)
    normalized_stem = normalize_key(stem)

    if normalized_stem in translations:
        return translations[normalized_stem] + suffix

    tokens = re.split(r"([_\-\s(),]+)", stem)
    translated_parts: list[str] = []

    for token in tokens:
        key = normalize_key(token)
        if not key:
            translated_parts.append(token)
            continue
        translated_parts.append(translations.get(key, DEFAULT_TRANSLATIONS.get(key, token)))

    translated = "".join(translated_parts)
    translated = translated.replace("_", " ")
    translated = re.sub(r"\s+", " ", translated).strip()
    translated = re.sub(r"\s+([),])", r"\1", translated)
    translated = re.sub(r"([(])\s+", r"\1", translated)
    return translated + suffix


def translated_relative_path(path: Path, root: Path, translations: dict[str, str]) -> str:
    relative_parts = path.relative_to(root).parts
    translated_parts = [translate_name(part, translations) for part in relative_parts]
    return str(Path(*translated_parts))


def iter_project_entries(root: Path, output_file: Path) -> list[Path]:
    entries = [
        path
        for path in root.rglob("*")
        if path.resolve() != output_file.resolve()
    ]
    return sorted(entries, key=lambda item: (str(item.parent).casefold(), item.name.casefold()))


def build_workbook(root: Path, output_file: Path, translations: dict[str, str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "nomes_traduzidos"

    headers = [
        "tipo",
        "nivel",
        "pasta_pai_original",
        "nome_original",
        "nome_ingles",
        "caminho_relativo_original",
        "caminho_relativo_ingles",
        "extensao",
        "tamanho_bytes",
        "ultima_modificacao",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for path in iter_project_entries(root, output_file):
        relative_path = path.relative_to(root)
        stat = path.stat()
        entry_type = "pasta" if path.is_dir() else "arquivo"
        parent = "" if path.parent == root else str(path.parent.relative_to(root))

        sheet.append(
            [
                entry_type,
                len(relative_path.parts),
                parent,
                path.name,
                translate_name(path.name, translations),
                str(relative_path),
                translated_relative_path(path, root, translations),
                "" if path.is_dir() else path.suffix,
                "" if path.is_dir() else stat.st_size,
                datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 70)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Cria um Excel com nomes originais de pastas, subpastas e arquivos e suas traduções para inglês."
    )
    parser.add_argument(
        "--root",
        default=".",
        help="Pasta raiz do projeto que será varrida. Padrão: pasta atual.",
    )
    parser.add_argument(
        "--output",
        default="nomes_pastas_arquivos_traduzidos.xlsx",
        help="Arquivo Excel de saída.",
    )
    parser.add_argument(
        "--translations",
        help="CSV opcional com colunas original,english para traduções manuais.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(args.root).resolve()
    output_file = Path(args.output).resolve()
    translation_file = Path(args.translations).resolve() if args.translations else None

    if not root.exists() or not root.is_dir():
        raise FileNotFoundError(f"Pasta raiz não encontrada: {root}")

    translations = load_translation_overrides(translation_file)
    build_workbook(root, output_file, translations)
    print(f"Excel criado: {output_file}")


if __name__ == "__main__":
    main()
